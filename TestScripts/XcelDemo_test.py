import openpyxl

# give the file location and read into one object
xl_sheet_obj = openpyxl.load_workbook("D:/PythonLearning2025/MockData/Xcel_Mockdata.xlsx")

# Read the active sheet into active_sheet_obj
active_sheet_obj = xl_sheet_obj.active

# Now read first cell of sheet
first_cell = active_sheet_obj.cell(row=1, column=1)
print(first_cell.value)

# Write data at 2nd cell in the first row
active_sheet_obj.cell(row=1, column=2).value = "Test Data at Row 1 column 2"
print(active_sheet_obj.cell(row=1, column=2).value)

# Get Max columns in the sheet
print("Max Columns:", active_sheet_obj.max_column)

# Get Max Rows in the sheet
print("Max rows:", active_sheet_obj.max_row)

# Print all values of the sheet
for row in range(1, (active_sheet_obj.max_row + 1)):  # To print the rows data
    if active_sheet_obj.cell(row=row, column=1).value == "shubham":
        print("data:")
        for column in range(1, (active_sheet_obj.max_column + 1)):  # To print the column data
            print(active_sheet_obj.cell(row=row, column=column).value)
